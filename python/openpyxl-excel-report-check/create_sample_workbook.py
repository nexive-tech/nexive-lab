from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_PATH = BASE_DIR / "samples" / "sample_invoice_report_check.xlsx"


def main() -> None:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "invoice"

    ws["A1"] = "請求書チェックサンプル"
    ws["A1"].font = Font(bold=True, size=14)

    header_rows = [
        ("A3", "請求番号", "B3", "INV-2026-0806-002"),
        ("A4", "取引先名", "B4", None),
        ("A5", "請求日", "B5", "2026-08-06"),
        ("A6", "請求金額", "B6", 120000),
    ]

    for label_cell, label, value_cell, value in header_rows:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True)
        ws[value_cell] = value

    headers = ["明細No", "品目コード", "作業内容", "数量", "単価", "金額"]
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center")

    rows = [
        [1, "SVC-001", "月次保守", 1, 60000, 60000],
        [2, "SVC-002", "追加調査", 1, 35000, 35000],
        [3, "SVC-003", None, 1, 20000, 20000],
        [4, "SVC-004", "交通費", 1, 5000, None],
    ]

    for row_index, row in enumerate(rows, start=10):
        for col_index, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index, value=value)

    for col_index, width in enumerate([10, 14, 24, 8, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width

    wb.save(OUTPUT_PATH)
    print(f"created: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
