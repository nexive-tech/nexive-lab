import argparse
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_PATH = BASE_DIR / "samples" / "sample_invoice_report_check.xlsx"
DEFAULT_OUTPUT_PATH = BASE_DIR / "check_results.csv"


@dataclass
class Issue:
    issue_type: str
    cell: str
    item: str
    value: Any
    message: str


def main() -> int:
    parser = argparse.ArgumentParser(description="Check blank cells and total mismatch in an Excel invoice report.")
    parser.add_argument("input", nargs="?", default=str(DEFAULT_INPUT_PATH), help="Path to an Excel file.")
    parser.add_argument("--out", default=str(DEFAULT_OUTPUT_PATH), help="Output CSV path.")
    args = parser.parse_args()

    issues = check_invoice_report(Path(args.input))
    write_issues(Path(args.out), issues)

    print(f"checked: {Path(args.input)}")
    print(f"issues: {len(issues)}")
    print(f"output: {Path(args.out)}")
    return 1 if issues else 0


def check_invoice_report(input_path: Path) -> list[Issue]:
    wb = load_workbook(input_path, data_only=True)
    ws = wb["invoice"]
    issues: list[Issue] = []

    required_header_cells = {
        "B3": "請求番号",
        "B4": "取引先名",
        "B5": "請求日",
        "B6": "請求金額",
    }

    for cell, item in required_header_cells.items():
        add_blank_issue(issues, cell, item, ws[cell].value)

    detail_total = 0
    for row in range(10, 14):
        row_no = ws.cell(row=row, column=1).value
        if is_blank(row_no):
            continue

        required_detail_cells = {
            f"B{row}": "品目コード",
            f"C{row}": "作業内容",
            f"D{row}": "数量",
            f"E{row}": "単価",
            f"F{row}": "金額",
        }

        for cell, item in required_detail_cells.items():
            add_blank_issue(issues, cell, f"明細{row_no}の{item}", ws[cell].value)

        amount = ws.cell(row=row, column=6).value
        if not is_blank(amount):
            detail_total += int(amount)

    billed_total = ws["B6"].value
    if not is_blank(billed_total) and int(billed_total) != detail_total:
        issues.append(
            Issue(
                issue_type="total_mismatch",
                cell="B6",
                item="請求金額",
                value=billed_total,
                message=f"請求金額 {billed_total} と明細合計 {detail_total} が一致しません",
            )
        )

    return issues


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def add_blank_issue(issues: list[Issue], cell: str, item: str, value: Any) -> None:
    if is_blank(value):
        issues.append(
            Issue(
                issue_type="blank",
                cell=cell,
                item=item,
                value=value,
                message=f"{item} が空欄です",
            )
        )


def write_issues(output_path: Path, issues: list[Issue]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["issue_type", "cell", "item", "value", "message"],
        )
        writer.writeheader()
        for issue in issues:
            writer.writerow(issue.__dict__)


if __name__ == "__main__":
    raise SystemExit(main())
