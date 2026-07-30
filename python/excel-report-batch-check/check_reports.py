import argparse
import csv
import json
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook


def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(description="Check Excel invoice reports in a folder.")
    parser.add_argument("input_dir", help="Folder that contains .xlsx files.")
    parser.add_argument("rules", help="JSON rule file.")
    parser.add_argument("--out", default="result.csv", help="Output CSV path.")
    args = parser.parse_args()

    rules = json.loads(Path(args.rules).read_text(encoding="utf-8"))
    errors = []

    for workbook_path in sorted(Path(args.input_dir).glob("*.xlsx")):
        errors.extend(check_workbook(workbook_path, rules))

    write_result(Path(args.out), errors)

    if errors:
        print(f"NG: {len(errors)}件の問題が見つかりました。結果: {args.out}")
        return 1

    print(f"OK: Excel帳票のチェックに成功しました。結果: {args.out}")
    return 0


def check_workbook(workbook_path, rules):
    errors = []
    workbook = load_workbook(workbook_path, data_only=True)
    sheet_name = rules["sheet_name"]

    if sheet_name not in workbook.sheetnames:
        return [error(workbook_path, "", f"シート「{sheet_name}」がありません")]

    sheet = workbook[sheet_name]
    check_required_cells(workbook_path, sheet, rules, errors)
    detail_total = check_detail_rows(workbook_path, sheet, rules, errors)
    check_totals(workbook_path, sheet, rules, detail_total, errors)

    return errors


def check_required_cells(workbook_path, sheet, rules, errors):
    for cell, label in rules["required_cells"].items():
        if is_blank(sheet[cell].value):
            errors.append(error(workbook_path, cell, f"{label}が空です"))


def check_detail_rows(workbook_path, sheet, rules, errors):
    detail = rules["detail"]
    total = Decimal("0")

    for row in range(detail["start_row"], detail["end_row"] + 1):
        item_cell = f'{detail["item_col"]}{row}'
        quantity_cell = f'{detail["quantity_col"]}{row}'
        unit_price_cell = f'{detail["unit_price_col"]}{row}'
        amount_cell = f'{detail["amount_col"]}{row}'

        item = sheet[item_cell].value
        quantity = sheet[quantity_cell].value
        unit_price = sheet[unit_price_cell].value
        amount = sheet[amount_cell].value

        if is_blank(item) and is_blank(quantity) and is_blank(unit_price) and is_blank(amount):
            continue

        for cell, value, label in [
            (item_cell, item, "品名"),
            (quantity_cell, quantity, "数量"),
            (unit_price_cell, unit_price, "単価"),
            (amount_cell, amount, "金額")
        ]:
            if is_blank(value):
                errors.append(error(workbook_path, cell, f"明細の{label}が空です"))

        quantity_value = to_decimal(quantity)
        unit_price_value = to_decimal(unit_price)
        amount_value = to_decimal(amount)

        if quantity_value is None:
            errors.append(error(workbook_path, quantity_cell, "数量が数値ではありません"))
        if unit_price_value is None:
            errors.append(error(workbook_path, unit_price_cell, "単価が数値ではありません"))
        if amount_value is None:
            errors.append(error(workbook_path, amount_cell, "金額が数値ではありません"))

        if quantity_value is None or unit_price_value is None or amount_value is None:
            continue

        expected_amount = quantity_value * unit_price_value
        if amount_value != expected_amount:
            errors.append(error(workbook_path, amount_cell, f"金額が数量×単価と一致しません。期待値: {expected_amount}"))

        total += amount_value

    return total


def check_totals(workbook_path, sheet, rules, detail_total, errors):
    totals = rules["totals"]
    subtotal = to_decimal(sheet[totals["subtotal_cell"]].value)
    tax = to_decimal(sheet[totals["tax_cell"]].value)
    total = to_decimal(sheet[totals["total_cell"]].value)
    expected_tax = yen(detail_total * Decimal(str(totals["tax_rate"])))
    expected_total = detail_total + expected_tax

    if subtotal != detail_total:
        errors.append(error(workbook_path, totals["subtotal_cell"], f"小計が明細合計と一致しません。期待値: {detail_total}"))

    if tax != expected_tax:
        errors.append(error(workbook_path, totals["tax_cell"], f"消費税が小計×税率と一致しません。期待値: {expected_tax}"))

    if total != expected_total:
        errors.append(error(workbook_path, totals["total_cell"], f"合計金額が小計+消費税と一致しません。期待値: {expected_total}"))


def write_result(path, errors):
    with path.open("w", newline="", encoding="utf-8-sig") as output:
        writer = csv.DictWriter(output, fieldnames=["file", "cell", "level", "message"])
        writer.writeheader()
        writer.writerows(errors)


def error(workbook_path, cell, message):
    return {
        "file": workbook_path.name,
        "cell": cell,
        "level": "error",
        "message": message
    }


def is_blank(value):
    return value is None or str(value).strip() == ""


def to_decimal(value):
    if is_blank(value):
        return None

    try:
        return Decimal(str(value))
    except Exception:
        return None


def yen(value):
    return value.quantize(Decimal("1"), rounding=ROUND_HALF_UP)


if __name__ == "__main__":
    raise SystemExit(main())
